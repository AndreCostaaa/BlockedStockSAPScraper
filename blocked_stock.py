from dataclasses import dataclass, asdict
from typing import Optional, Any
from dacite import from_dict
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Color,PatternFill, NamedStyle
from openpyxl.utils.cell import get_column_letter

from datetime import datetime
import pickle

from config import ARCHIVED_BIN_PATH, STOCK_BIN_PATH


DATE_TIME_FORMAT = "%Y/%m/%d %H:%M:%S"
RRP_LOCATIONS = {"NX42", "NX72", "SF42"}
CC_LOCATIONS = {"NT01", "NT72"}


def title_to_snake(value:str):
    return value.lower().replace(' ', '_')

def snake_to_title(value:str):
    return value.replace('_', ' ').title()

def get_date_now():
    return datetime.now().strftime(DATE_TIME_FORMAT)

@dataclass
class NCData:
    batch_sap:Optional[str]
    item_code:Optional[str]
    status_nc:Optional[str]
    vendor_name:Optional[str]
    origin:Optional[str]
    category:Optional[str]
    purchase_order:Optional[str]
    delivery_date:Optional[datetime]
    lu:Optional[str]
    vendor_id:Optional[str]
    responsable_nc:Optional[str]
    created_by:Optional[str]
    created:Optional[datetime]
    dim_derogation_conv:Optional[str]
    quantity:Optional[str]
    uom:Optional[str]
    commentaire:Optional[str]
    code_raison_de_blocage:Optional[str]
    decision_mqa:Optional[str]
    type_nc:Optional[str]
    description_nc:Optional[str]
    impact_on_brand_image:Optional[bool]
    impact_on_product_perfomance:Optional[bool]
    impact_on_regulatory:Optional[bool]
    impact_on_product_identification:Optional[bool]
    impact_on_machine_efficiency:Optional[bool]
    recurrent_nc:Optional[bool]
    risk_evaluation:Optional[str]
    vendor_batch:Optional[str]
    approver:Optional[str]
    capa_impl_data:Optional[datetime] #CAPA-Impl. Date
    destruction_at_vendors_cost:Optional[str] #Destruction at vendor's cost?
    traitement_nc:Optional[str] #Traitement NC (À remplir si pas de NCR faite)
    total_cost_over_200:Optional[bool] #Total cost >200CHF
    nc_nr:Optional[str]
    ncr_number:Optional[str]
    opening_ncr_date:Optional[datetime]
    stock_exit_code:Optional[str]
    number_refresh_capa:Optional[int]
    data_capa_refresh:Optional[datetime]
    item_type:Optional[str]
    path:Optional[str]
@dataclass
class StockExitData:
    stock_exit_code:Optional[str]
    nc_nr:Optional[str]
    batch_sap:Optional[str]
    item_code:Optional[str]
    purchase_order:Optional[str]
    quantity:Optional[str]
    uom:Optional[str]
    vendor_batch_id:Optional[str]
    ncr_number:Optional[str]
    initial_location:Optional[str]
    stock_exit_type:Optional[str]
    cost_center:Optional[str]
    wbs:Optional[str]
    action_dim:Optional[datetime]
    reception_entrepot:Optional[str]
    odn_number:Optional[str]
    odn_comment:Optional[str]
    stock_exit_sap:Optional[datetime]
    odn_completed:Optional[datetime]
    dim_action:Optional[str]
    action_wh:Optional[str]
    action_warehouse_completed:Optional[str]
    information_transporteur:Optional[str]
    a_traiter_par_chavornay:Optional[bool]
    item_type:Optional[str]
    path:Optional[str]

@dataclass
class BlockedStockLine:
    material:str
    batch:str
    material_description:str
    material_type:str
    plant: str
    storage_location:str
    base_unit_of_measure:str
    blocked:float
    value_blockedstock:float
    currency:str
    date_created:str
    date_archived:Optional[str]
    done:bool
    type:str
    resp:Optional[str]
    provision:Optional[str]
    commentaire:Optional[str]
    nc_number_match:int
    ncf:Optional[str]
    ncr:Optional[str]
    others_ncf:Optional[str]
    others_ncr:Optional[str]
    step_ncf:Optional[str]
    stock_exit_type:Optional[str]
    vendor_id:Optional[str]
    reason_code:Optional[str]
    vendor_name:Optional[str]
    destruction_at_vendors_cost:Optional[str]

class ExcelParser:
    @staticmethod
    def parse_data(sheet:Worksheet) -> dict[str,str]:
        data = []
        for i, row in enumerate(sheet.rows):
            if i == 0:
                continue
            dic = {}
            for j, cell in enumerate(row):
                key = sheet.cell(row = 1, column= j + 1).value
                if not key:
                    continue
                if cell.fill.start_color.index == 13:
                    break
                dic[title_to_snake(key)] = cell.value
            if dic:
                data.append(dic)
        return data

class StaticData:
    def __init__(self, sheet:Worksheet, type: NCData | StockExitData, keys_translator_map: dict[str,str]):
        self.type = type
        self.sheet = sheet
        self.data: dict[str, list[type]] = {}
        self.translator_map = keys_translator_map
    
    def translate_keys(self, obj:dict[str, str]):
        for src_key, dst_key in self.translator_map.items():
            obj[dst_key] = obj[src_key]
        
    @staticmethod
    def is_valid_obj(obj):
        return obj["item_code"] and obj["batch_sap"]

    @staticmethod
    def get_keys_missing(obj):
        keys_missing = []
        for key in ["item_code", "batch_sap"]:
            if not obj[key]:
                keys_missing.append(key)

        return keys_missing
    
    def load_data(self):
        parsed_data = ExcelParser.parse_data(self.sheet)
        self.data = {}
        for i, obj in enumerate(parsed_data):
            if not self.is_valid_obj(obj):
                print(f"[WARNING] [{self.sheet.title}] Manque {[snake_to_title(key) for key in self.get_keys_missing(obj)]} ligne: {i+2}")
                continue
            self.translate_keys(obj)

            obj_id = hash(obj["item_code"] + obj["batch_sap"])
            obj = from_dict(self.type, obj)
            if self.data.get(obj_id):
                self.data[obj_id].append(obj)
            else:
                self.data[obj_id] = [obj]


class Ncs:
    def __init__(self, sheet:Worksheet):
        self.sheet = sheet
        self.nc_data: dict[str, list[NCData]] = {}

    @staticmethod
    def translate_keys(obj:dict[str, Any]):
        translator_map = {  
            "vendor_id:vendor_name_(linked_to_item)" :"vendor_name",
            "destruction_at_vendor's_cost?":  "destruction_at_vendors_cost",
            "capa-impl._date": "capa_impl_data",
            "traitement_nc_(à_remplir_si_pas_de_ncr_faite)": "traitement_nc",
            "total_cost_>200chf": "total_cost_over_200"
        }

        for src_key, dst_key in translator_map.items():
            obj[dst_key] = obj[src_key]

    @staticmethod
    def is_valid_obj(obj):
        ret = True
        keys_missing = []
        for key in ["item_code", "batch_sap"]:
            if not obj[key]:
                ret = False
                keys_missing.append(key)

        return ret, keys_missing
    
    def load_data(self):
        parsed_data = ExcelParser.parse_data(self.sheet)
        self.nc_data = {}
        for i, obj in enumerate(parsed_data):
            valid, missing_keys = self.is_valid_obj(obj)
            if not valid:
                print(f"[WARNING] [{self.sheet.title}] Missing {[snake_to_title(key) for key in missing_keys]} on line {i+2}. Skipping...")
                continue
            self.translate_keys(obj)

            obj_id = hash(obj["item_code"] + obj["batch_sap"])
            obj = from_dict(NCData, obj)
            if self.nc_data.get(obj_id):
                self.nc_data[obj_id].append(obj)
            else:
                self.nc_data[obj_id] = [obj]


class StockExitNew:
    def __init__(self, sheet:Worksheet):
        self.sheet = sheet
        self.stock_exit_data:dict[int, list[StockExitData]]

    @staticmethod
    def is_valid_obj(obj):
        ret = True
        keys_missing = []
        for key in ["item_code", "batch_sap"]:
            if not obj[key]:
                ret = False
                keys_missing.append(key)

        return ret, keys_missing
    
    def load_data(self):
        parsed_data = ExcelParser.parse_data(self.sheet)
        self.stock_exit_data = {}
        for i, obj in enumerate(parsed_data):
            valid, missing_keys = self.is_valid_obj(obj)
            if not valid:
                print(f"[WARNING] [{self.sheet.title}] Missing {[snake_to_title(key) for key in missing_keys]} on line {i+2}. Skipping...")
                continue

            obj_id = hash(obj["item_code"] + obj["batch_sap"])
            obj = from_dict(StockExitData, obj)
            if self.stock_exit_data.get(obj_id):
                self.stock_exit_data[obj_id].append(obj)
            else:
                self.stock_exit_data[obj_id] = [obj]


class BlockedStock:
    def __init__(self, sheet:Worksheet):
        self.sheet = sheet
        self.blocked_stock: dict[int, list[BlockedStockLine]] = {}
        #self.load_data()

    def load_data(self, ncs:StaticData = None, stock_exit:StaticData = None):
        parsed_data = ExcelParser.parse_data(self.sheet)
        self.blocked_stock = {}
        for obj in parsed_data:
            obj["type"] = ""
            if obj["storage_location"] in RRP_LOCATIONS:
                obj["type"] = "RRP"
            elif obj["storage_location"] in CC_LOCATIONS:
                obj["type"] = "CC"
            
            is_done = obj.get("done")
            if is_done is None:
                is_done = False
            obj["date_created"] = obj.get("date_created", get_date_now())
            date_archived = obj.get("date_archived", "")

            obj["date_archived"] = date_archived
            obj["done"] = is_done

            obj_id = hash(obj["material"] + obj["batch"])
            if ncs:
                self.add_nc_data_to_obj(obj, ncs.data.get(obj_id, []))
            if stock_exit:
                self.add_stock_exit_data_to_obj(obj, stock_exit.data.get(obj_id, []))

            obj = from_dict(BlockedStockLine,obj)

            if self.blocked_stock.get(obj_id):
                self.blocked_stock[obj_id].append(obj)
            else:
                self.blocked_stock[obj_id] = [obj]

    def add(self, key: int, obj:BlockedStockLine) -> None:
        if key in self.blocked_stock:
            self.blocked_stock[key].append(obj)
        else:
            self.blocked_stock[key] = [obj]
        
    def check_and_transfer_done_objs(self, dst, f = None) -> None:
        keys_to_pop = []
        for key, lst in self.blocked_stock.items():
            for i, obj in enumerate(sorted(lst,key=lambda x: x.batch,reverse=True)):
                if obj.done:
                    if f:
                        f(obj)
                    dst.add(key,obj)
                    lst.pop(i)
            if len(lst) == 0:
                keys_to_pop.append(key)
                    
        for key in keys_to_pop:
            del self.blocked_stock[key]

    def add_nc_data_to_obj(self, obj, nc_lst:list[NCData]):
        nc_args = ["ncf", "ncr", "step_ncf", "vendor_id", "vendor_name", "others_ncf", "others_ncr"]
        obj["nc_number_match"] = len(nc_lst)
        if not nc_lst:
            for arg in nc_args:
                obj[arg] = None    
            return
        
        obj["ncf"] = nc_lst[0].nc_nr
        obj["ncr"] = nc_lst[0].ncr_number
        obj["step_ncf"] = nc_lst[0].status_nc
        obj["vendor_id"] = nc_lst[0].vendor_id
        obj["vendor_name"] = nc_lst[0].vendor_name
        obj["resp"] = nc_lst[0].responsable_nc
        obj["destruction_at_vendors_cost"] = nc_lst[0].destruction_at_vendors_cost
        obj["others_ncf"] = ""
        obj["others_ncr"] = ""
        for i, o in enumerate(nc_lst):
            if i == 0:
                continue
            if o.nc_nr:
                obj["others_ncf"] += o.nc_nr + " "
            if o.ncr_number:
                obj["others_ncr"] += o.ncr_number + " "
        

    def add_stock_exit_data_to_obj(self, obj, stock_exit_lst:list[StockExitData]):
        if not stock_exit_lst:
            return
        obj["stock_exit_type"] = stock_exit_lst[0].stock_exit_type
        
    def write_data(self):

        self.sheet.delete_rows(1,self.sheet.max_row)
        if not self.blocked_stock:
            return
        keys = list(self.blocked_stock.keys())
        j = 1

        for key in asdict(self.blocked_stock[keys[0]][0]):
            self.sheet.cell(row=1, column=j).value = snake_to_title(key)
            j += 1

        i = 2
        for key, lst in self.blocked_stock.items():
            for obj in lst:
                dict_obj = asdict(obj)
                j = 1
                for key, val in dict_obj.items():
                    self.sheet.cell(row=i, column=j).value = val
                    j += 1
                i += 1

    def add_data_validation(self, data_validation:dict[str, DataValidation]):
        if not data_validation:
            return
        
        for _, obj in data_validation.items():
            self.sheet.add_data_validation(obj)

        mr = self.sheet.max_row
        if mr < 2:
            return
        for i in range(1,self.sheet.max_column + 1):
            cell = self.sheet.cell(1,i)
            key = title_to_snake(cell.value)
            if key in data_validation:
                column_letter = get_column_letter(i)
                whole_column = f'{column_letter}2:{column_letter}{mr}' #eg: 'B2:B233' 
                data_validation[key].add(whole_column)


    def make_pretty(self):

        title_font = Font(color="00000000", bold=True, sz=11)
        for cell in self.sheet["1:1"]:
            cell.font= title_font
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        self.sheet.auto_filter.ref = self.sheet.dimensions
        for i, row in enumerate(self.sheet):
            for cell in row:
                cell.alignment=Alignment(wrap_text=True, horizontal="center", vertical="center")
            if i%2 == 0:
                for cell in row:
                    cell.fill = PatternFill(patternType="solid", fill_type="solid", fgColor=Color("C4C4C4"))
        for i in range(1, self.sheet.max_column + 1):
            max_len = len(str(self.sheet.cell(row=1, column=i).value))
            for j in range(2,self.sheet.max_row):
                if not str(self.sheet.cell(row=j, column=i).value):
                    continue
                curr_len = len(str(self.sheet.cell(row=j, column=i).value))
                if curr_len > max_len:
                    max_len = curr_len
            width = (max_len + 2) * 1.2
            self.sheet.column_dimensions[xl.utils.get_column_letter(i)].width = width   

        



class DbHandler:

    def __init__(self, file_path:str):
        self.file_path = file_path
        self.workbook = xl.load_workbook(file_path)
        self.current = BlockedStock(self.workbook["Current Blocked Stock"])
        self.disappeared = BlockedStock(self.workbook["Gone from SAP"])
        self.archive = BlockedStock(self.workbook["Archive"])
        self.nc_data = StaticData(self.workbook["NCs"], NCData, {  
            "vendor_id:vendor_name_(linked_to_item)" :"vendor_name",
            "destruction_at_vendor's_cost?":  "destruction_at_vendors_cost",
            "capa-impl._date": "capa_impl_data",
            "traitement_nc_(à_remplir_si_pas_de_ncr_faite)": "traitement_nc",
            "total_cost_>200chf": "total_cost_over_200"
        })

        self.stock_exit_new = StaticData(self.workbook["Stock Exit New"], StockExitData, {})

        self.nc_data.load_data()
        self.stock_exit_new.load_data()
        self.full_stock: dict[int, list[BlockedStockLine]]
        for stock in [self.current, self.disappeared, self.archive]:
            stock.load_data(self.nc_data, self.stock_exit_new)
        
    

    def handle_differences(self, new_blocked_stock:BlockedStock):
        
        static_args = [
            "date_created",
            "type",
            "resp",
            "provision",
            "done",
            "commentaire",
            "ncf",
            "ncr",
            "step_ncf",
            "stock_exit_type",
            "vendor_id",
            "reason_code",
            "vendor_name",
            "destruction_at_vendors_cost"
        ]

        #Current stock + Disappeared -> Archive/Disappeared
        for key, lst in self.current.blocked_stock.items():
            #stock gone
            if key not in new_blocked_stock.blocked_stock:
                for obj in lst:
                    if obj.done:
                        obj.date_archived = get_date_now()
                        self.archive.add(key, obj)
                    else:
                        self.disappeared.add(key,obj)                    
        #New Stock -> Current Stock
        for key, lst in new_blocked_stock.blocked_stock.items():
            
            for i, obj in enumerate(lst):
                if key not in self.current.blocked_stock:
                    continue
                for arg in static_args:
                    obj.__setattr__(arg, self.current.blocked_stock[key][i].__getattribute__(arg))

        self.current.blocked_stock = new_blocked_stock.blocked_stock

        def add_date_archived(obj:BlockedStockLine):
            obj.date_archived = get_date_now()

        self.disappeared.check_and_transfer_done_objs(self.archive, add_date_archived)
           
        


    def save(self, file_path:str = ""):
        
        data_validation_map = {"provision": DataValidation(type="list", showDropDown=False,allow_blank=True,formula1= f'"{",".join(["N", "TBD", "PMP SA", "PMP SA/Affiliate", "Affiliate", "Vendor", "Supply Chain"])}"'),
                               "resp": DataValidation(type="list", showDropDown=False, allow_blank=True, formula1= f'"{",".join(["Gilles","Serge", "Quentin", "Ariane", "Raphael B."])}"'),
                               "done": DataValidation(type="list", showDropDown=False, allow_blank=False, formula1='"FALSE,TRUE"')}
        
        for sheet in [self.current, self.disappeared, self.archive]:
            sheet.write_data()
            sheet.add_data_validation(data_validation_map)
            sheet.make_pretty()
        #dv = DataValidation(type="list", showDropDown=False, allow_blank=False,formula1= f'"{",".join(["Gilles","Serge", "Quentin", "Ariane", "Raphael B."])}"')
        #self.current.sheet.add_data_validation(dv)

        #dv.add("N1:N1048576")
        #print("N4" in dv)
        pickle.dump(self.current.blocked_stock, open(STOCK_BIN_PATH, "wb"),protocol=-1)
        pickle.dump(self.archive.blocked_stock, open(ARCHIVED_BIN_PATH, "wb"),protocol=-1)

        if not file_path:
            file_path = self.file_path
            
        self.workbook.save(self.file_path)


def get_current_stock(data_path, ncs, stock_exit):
    file = xl.load_workbook(filename=data_path, read_only=True)
    sheet = file["Sheet1"]
    stock = BlockedStock(sheet)
    stock.load_data(ncs, stock_exit)
    return stock