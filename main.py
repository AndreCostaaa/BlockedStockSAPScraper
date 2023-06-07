import os
import openpyxl as xl
from blocked_stock import BlockedStock, BlockedStockLine, DbHandler, get_current_stock
from pprint import pprint
from dataclasses import asdict
from config import load_config

from data_fetcher import fetch_data_wait


def copy_data(source_data, dest_file, sheet_name, dest_start_row, dest_start_column):

    dest_sheet = dest_file[sheet_name]
    mc = dest_sheet.max_column
    over = False
    i = dest_start_row
    for obj in source_data:
        j = dest_start_column
        for _, value in obj.items():
            dest_sheet.cell(row=i, column=j).value = value
            j += 1
        i += 1

    while not over:
        for j in range(dest_start_column, mc + 1):
            val = dest_sheet.cell(row=i, column=j).value
            if not val:
                over = True
                break
            dest_sheet.cell(row=i, column=j).value = ""
        i += 1

def close_windows():

    os.system(f"taskkill /f /im saplogon.exe")
    os.system(f"taskkill /f /im excel.exe")
    
import time
def main():

    config = load_config()
    
    close_windows()
    time.sleep(1)
    data_path = fetch_data_wait(config.sap_path)    
    print("Calculating current stock. Please wait, this may take a while")
    curr_blocked_stock = get_current_stock(r"C:\temp\EXPORT.xlsx" )

    print("Calculating differences")
    db = DbHandler(config.xlsx_db_path)
    db.handle_differences(curr_blocked_stock)
    print("Exporting data")
    db.save()
    close_windows()

if __name__ == "__main__":
    main()