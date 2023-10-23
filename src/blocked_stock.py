from dataclasses import asdict
import os

from dacite import from_dict
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from openpyxl.styles import Font, Alignment, Color,PatternFill
from openpyxl.utils.cell import get_column_letter
import json
from datetime import datetime

from data_classes import NC_TRANSLATOR_MAP, BlockedStockLine, NCData, StockExitData


DATE_TIME_FORMAT = "%Y/%m/%d %H:%M:%S"

RRP_LOCATIONS = {"NX42", "NX72", "SF42"}
CC_LOCATIONS = {"NT01", "NT72"}

def parse_html(html:str):
    import re
    return re.sub("<[^>]*>", "", html)

def title_to_snake(value:str):
    return value.lower().replace(' ', '_')

def snake_to_title(value:str):
    return value.replace('_', ' ').title()

def get_date_now():
    return datetime.now().strftime(DATE_TIME_FORMAT)



class ExcelHelper:
    @staticmethod
    def parse_data(sheet:Worksheet) -> dict[str,str]:
        data = []
        for i, row in enumerate(sheet.rows):
            if i == 0:
                continue
            dic = {}
            for j, cell in enumerate(row):
                key = sheet.cell(row = 1, column= j + 1).value
                if not key:
                    continue
                if cell.fill.start_color.index == 13:
                    break
                value = cell.value
                
                if type(value) is str:
                    if "<div" in cell.value:
                        value = parse_html(value)
                    
                dic[title_to_snake(key)] = cell.value
            if dic:
                data.append(dic)
        return data
    @staticmethod
    def write_data(sheet:Worksheet, data:dict):
        
        sheet.delete_rows(1,sheet.max_row)
        if not data:
            return
        keys = list(data.keys())
        j = 1

        for key in asdict(data[keys[0]][0]):
            sheet.cell(row=1, column=j).value = snake_to_title(key)
            j += 1

        i = 2
        for key, lst in data.items():
            for obj in lst:
                dict_obj = asdict(obj)
                j = 1
                for key, val in dict_obj.items():
                    sheet.cell(row=i, column=j).value = val
                    j += 1
                i += 1
    @staticmethod
    def make_pretty(sheet):
        title_font = Font(color="00000000", bold=True, sz=11)
        for cell in sheet["1:1"]:
            cell.font= title_font
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        sheet.auto_filter.ref = sheet.dimensions
        for i, row in enumerate(sheet):
            for cell in row:
                cell.alignment=Alignment(wrap_text=True, horizontal="center", vertical="center")
            if i%2 == 0:
                for cell in row:
                    cell.fill = PatternFill(patternType="solid", fill_type="solid", fgColor=Color("C4C4C4"))
        for i in range(1, sheet.max_column + 1):
            max_len = len(str(sheet.cell(row=1, column=i).value))
            for j in range(2,sheet.max_row):
                if not str(sheet.cell(row=j, column=i).value):
                    continue
                curr_len = len(str(sheet.cell(row=j, column=i).value))
                if curr_len > max_len:
                    max_len = curr_len
            width = (max_len + 2) * 1.2
            sheet.column_dimensions[xl.utils.get_column_letter(i)].width = width   

class StaticData:
    def __init__(self, src_sheet:Worksheet, dest_sheet:Worksheet, type: NCData | StockExitData, keys_translator_map: dict[str,str]):
        self.type = type
        self.source_sheet = src_sheet
        self.dest_sheet = dest_sheet
        self.data: dict[str, list[type]] = {}
        self.translator_map = keys_translator_map
    
    def translate_keys(self, obj:dict[str, str]):
        for src_key, dst_key in self.translator_map.items():
            obj[dst_key] = obj[src_key]
        
    @staticmethod
    def is_valid_obj(obj):
        return obj["item_code"] and obj["batch_sap"]

    @staticmethod
    def get_keys_missing(obj):
        keys_missing = []
        for key in ["item_code", "batch_sap"]:
            if not obj[key]:
                keys_missing.append(key)

        return keys_missing
    
    def load_data(self):
        parsed_data = ExcelHelper.parse_data(self.source_sheet)
        self.data = {}
        for i, obj in enumerate(parsed_data):
            if not self.is_valid_obj(obj):
                print(f"[WARNING] [{self.source_sheet.title}] Manque {[snake_to_title(key) for key in self.get_keys_missing(obj)]} ligne: {i+2}")
                continue
            self.translate_keys(obj)

            obj_id = hash(obj["item_code"] + obj["batch_sap"])

            obj = from_dict(self.type, obj)
            if self.data.get(obj_id):
                self.data[obj_id].append(obj)
            else:
                self.data[obj_id] = [obj]


    def write_data(self):
        ExcelHelper.write_data(self.dest_sheet, self.data)

    def make_pretty(self):
        ExcelHelper.make_pretty(self.dest_sheet)
        

class BlockedStock:
    def __init__(self, sheet:Worksheet):
        self.sheet = sheet
        self.blocked_stock: dict[int, list[BlockedStockLine]] = {}
        #self.load_data()

    def load_data(self, ncs:StaticData, stock_exit:StaticData, locations:dict[str, list]):
        parsed_data = ExcelHelper.parse_data(self.sheet)
        self.blocked_stock = {}
        for obj in parsed_data:
            
            
            if not obj.get("cc_rrp"):
                obj["cc_rrp"] = ""
                if obj["storage_location"] in locations["rrp"]:
                    obj["cc_rrp"] = "RRP"
                elif obj["storage_location"] in locations["cc"]:
                    obj["cc_rrp"] = "CC"
                
            
            is_done = obj.get("done")
            if is_done is None:
                is_done = False
            
            if isinstance(is_done,str) and is_done.lower() == "done":
                is_done = True
            obj["date_created"] = obj.get("date_created", get_date_now())
            date_archived = obj.get("date_archived", "")

            obj["date_archived"] = date_archived
            obj["done"] = is_done
            if not obj.get("type"):
                obj["type"] = ""

            obj_id = hash(obj["material"] + obj["batch"])
            if ncs:
                self.add_nc_data_to_obj(obj, ncs.data.get(obj_id, []))
            if stock_exit:
                self.add_stock_exit_data_to_obj(obj, stock_exit.data.get(obj_id, []))

            obj = from_dict(BlockedStockLine,obj)

            if self.blocked_stock.get(obj_id):
                self.blocked_stock[obj_id].append(obj)
            else:
                self.blocked_stock[obj_id] = [obj]

    def add(self, key: int, obj:BlockedStockLine) -> None:
        if key in self.blocked_stock:
            self.blocked_stock[key].append(obj)
        else:
            self.blocked_stock[key] = [obj]
        
    def check_and_transfer_done_objs(self, dst, f = None) -> None:
        keys_to_pop = []
        for key, lst in self.blocked_stock.items():
            indexes_to_pop = []
            for i, obj in enumerate(lst):
                if obj.done:
                    if f:
                        f(obj)
                    dst.add(key,obj)
                    indexes_to_pop.append(i)
                    
            for i in sorted(indexes_to_pop, reverse=True):
                lst.pop(i)

            if len(lst) == 0:
                keys_to_pop.append(key)
                    
        for key in keys_to_pop:
            del self.blocked_stock[key]

    def add_nc_data_to_obj(self, obj, nc_lst:list[NCData]):
        nc_args = ["ncf", "ncr", "step_ncf", "vendor_id", "vendor_name", "others_ncf", "others_ncr"]
        obj["nc_number_match"] = len(nc_lst)
        if not nc_lst:
            for arg in nc_args:
                obj[arg] = None    
            return
        
        obj["ncf"] = nc_lst[0].nc_number
        obj["ncr"] = nc_lst[0].ncr_number
        obj["step_ncf"] = nc_lst[0].status_nc
        obj["vendor_id"] = nc_lst[0].vendor_id
        obj["vendor_name"] = nc_lst[0].vendor_name
        obj["resp"] = nc_lst[0].responsable_nc
        obj["destruction_at_vendors_cost"] = nc_lst[0].destruction_at_vendors_cost
        obj["others_ncf"] = ""
        obj["others_ncr"] = ""
        for i, o in enumerate(nc_lst):
            if i == 0:
                continue
            if o.nc_number:
                obj["others_ncf"] += o.nc_number + " "
            if o.ncr_number:
                obj["others_ncr"] += o.ncr_number + " "
        

    def add_stock_exit_data_to_obj(self, obj, stock_exit_lst:list[StockExitData]):
        if not stock_exit_lst:
            return
        obj["stock_exit_type"] = stock_exit_lst[0].stock_exit_type
        
    def write_data(self):
        ExcelHelper.write_data(self.sheet, self.blocked_stock)
    

    def add_data_validation(self, data_validation:dict[str, DataValidation]):
        if not data_validation:
            return
        #remove current data validation
        self.sheet.data_validations = DataValidationList()

        for _, obj in data_validation.items():
            self.sheet.add_data_validation(obj)

        mr = self.sheet.max_row
        if mr < 2:
            return
        for i in range(1,self.sheet.max_column + 1):
            title_cell = self.sheet.cell(1,i)
            key = title_to_snake(title_cell.value)
            if key in data_validation:
                column_letter = get_column_letter(i)
               
                whole_column = f'{column_letter}2:{column_letter}{mr}' #eg: 'B2:B233' 
                data_validation[key].add(whole_column)


    def make_pretty(self):
        ExcelHelper.make_pretty(self.sheet)
        


class DbHandler:

    def __init__(self, file_path:str):
        self.sharepoint_file_path = r"C:\Users\acosta18\Philip Morris International\Quality Neuchatel - Documents\31_Quality_Shared\04_MQA\Analyse blocked stock\Blocked_Stock\BlockedStock IN TEST\SharePointData.xlsx"
        print(f"[DB Handler] Loading ${file_path}")
        self.file_path = file_path
        self.workbook = xl.load_workbook(file_path)
        print(f"[DB Handler] Loading 'SharePointData.xlsx'")
        self.external_data_workbook = xl.load_workbook(os.path.join(os.path.dirname(self.file_path), self.sharepoint_file_path))
        self.current = BlockedStock(self.workbook["Current Blocked Stock"])
        self.disappeared = BlockedStock(self.workbook["Gone from SAP"])
        self.archive = BlockedStock(self.workbook["Archive"])
        self.nc_data = StaticData(self.external_data_workbook["NCs"],self.workbook["NCs"], NCData, NC_TRANSLATOR_MAP)

        self.stock_exit_new = StaticData(self.external_data_workbook["Stock Exit New"], self.workbook["Stock Exit New"], StockExitData, {})

        self.full_stock: dict[int, list[BlockedStockLine]]

        self.nc_data.load_data()
        self.stock_exit_new.load_data()
        print(f"[DB Handler] Loading 'locations.json'")
        self.locations = json.load(open(os.path.join(os.path.dirname(self.file_path), "locations.json")))
        print(f"[DB Handler] Found the following locations")
        for key, obj in self.locations.items():
            print(f"\t{snake_to_title(key)} - {obj}")

        print(f"[DB Handler] Loading 'data_validation.json'")
        data_validation =  json.load(open(os.path.join(os.path.dirname(self.file_path), "data_validation.json")))
        print(f"[DB Handler] Found data validation for the following columns")
        for key, obj in data_validation.items():
            print(f"\t{snake_to_title(key)} - {obj}")

        self.data_validation_map = {key: DataValidation(type="list", showDropDown=False,allow_blank=True,formula1= f'"{",".join(obj)}"') for key, obj in data_validation.items() }

        for src in [self.current, self.disappeared, self.archive]:
            src.load_data(self.nc_data, self.stock_exit_new, self.locations)
        


    def handle_differences(self, new_blocked_stock:BlockedStock):
        
        static_args = [
            "date_created",
            "type",
            "resp",
            "provision",
            "done",
            "commentaire",
            "ncf",
            "ncr",
            "step_ncf",
            "stock_exit_type",
            "vendor_id",
            "reason_code",
            "vendor_name",
            "destruction_at_vendors_cost"
        ]

        #Current stock + Disappeared -> Archive/Disappeared
        for key, lst in self.current.blocked_stock.items():
            #stock gone
            if key not in new_blocked_stock.blocked_stock:
                for obj in lst:
                    if obj.done:
                        obj.date_archived = get_date_now()
                        self.archive.add(key, obj)
                    else:
                        self.disappeared.add(key,obj)                    
        #New Stock -> Current Stock
        for key, lst in new_blocked_stock.blocked_stock.items():
            
            for i, obj in enumerate(lst):
                if key not in self.current.blocked_stock:
                    continue
                for arg in static_args:
                    if i >= len(self.current.blocked_stock[key]):
                        break
                    obj.__setattr__(arg, self.current.blocked_stock[key][i].__getattribute__(arg))

        self.current.blocked_stock = new_blocked_stock.blocked_stock

        def add_date_archived(obj:BlockedStockLine):
            obj.date_archived = get_date_now()

        self.disappeared.check_and_transfer_done_objs(self.archive, add_date_archived)
           
        


    def save(self, file_path:str = ""):
        
        print("[DB Handler] Writing Data...")
        for sheet in [self.current, self.disappeared, self.archive]:
            sheet.write_data()
            sheet.add_data_validation(self.data_validation_map)
            sheet.make_pretty()
        print("[DB Handler] Finishing up...")
        for external_data in [self.nc_data, self.stock_exit_new]:
            external_data.write_data()
            external_data.make_pretty()
        #dv = DataValidation(type="list", showDropDown=False, allow_blank=False,formula1= f'"{",".join(["Gilles","Serge", "Quentin", "Ariane", "Raphael B."])}"')
        #self.current.sheet.add_data_validation(dv)

        #dv.add("N1:N1048576")
        #print("N4" in dv)
        #pickle.dump(self.current.blocked_stock, open(STOCK_BIN_PATH, "wb"),protocol=-1)
        #pickle.dump(self.archive.blocked_stock, open(ARCHIVED_BIN_PATH, "wb"),protocol=-1)

        if not file_path:
            file_path = self.file_path
            
        self.workbook.save(self.file_path)


def get_current_stock(data_path, db:DbHandler):
    file = xl.load_workbook(filename=data_path, read_only=True)
    sheet = file["Sheet1"]
    stock = BlockedStock(sheet)
    stock.load_data(db.nc_data, db.stock_exit_new,db.locations)
    return stock