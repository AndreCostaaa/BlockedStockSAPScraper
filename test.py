from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
def main():


    # Create the workbook and worksheet we'll be working with
    wb = Workbook()
    ws = wb.active

    # Create a data-validation object with list validation
    dv = DataValidation(type="list", formula1= f'"{",".join(["Gilles","Serge", "Quentin", "Ariane", "Raphael B."])}"', allow_blank=True)

    # Optionally set a custom error message
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'

    # Optionally set a custom prompt message
    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'List Selection'

    # Add the data-validation object to the worksheet
    ws.add_data_validation(dv)
     # Create some cells, and add them to the data-validation object
    c1 = ws["A1"]
    c1.value = "Dog"
    dv.add(c1)
    c2 = ws["A2"]
    c2.value = "An invalid value"
    dv.add(c2)

    # Or, apply the validation to a range of cells
    dv.add('B2:B1048576') # This is the same as for the whole of column B
    wb.save("test.xlsx")


if __name__ == "__main__":
    main()